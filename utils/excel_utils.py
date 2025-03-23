import openpyxl
import os
from tkinter import messagebox, filedialog
from datetime import datetime
from openpyxl.styles import PatternFill

# --- Cargar maestro de exámenes ---
def cargar_listado_examenes_desde_excel():
    examenes = []
    try:
        ruta_archivo = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'maestro_examenes.xlsx')
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            codigo = row[0]
            descripcion = row[1]
            if codigo and descripcion:
                examenes.append((codigo, descripcion))

        return examenes
    except Exception as e:
        print(f"Error al cargar exámenes: {str(e)}")
        return []

# --- Generar listado Excel ---
def generar_listado_excel(pacientes):
    from tkinter import Tk
    root = Tk()
    root.withdraw()
    carpeta = filedialog.askdirectory(title="Selecciona carpeta destino")
    if not carpeta:
        messagebox.showinfo("Info", "Operación cancelada")
        return

    try:
        fecha_actual = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
        nombre_archivo = f"listado_derivacion_{fecha_actual}.xlsx"
        archivo_excel = os.path.join(carpeta, nombre_archivo)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Listado de Derivación"

        sheet.cell(row=2, column=1, value="Fecha Actual")
        sheet.cell(row=2, column=2, value=datetime.now().strftime('%d/%m/%Y'))
        sheet.cell(row=4, column=1, value="Código de Cliente")
        sheet.cell(row=4, column=2, value="474")

        headers = ["Código Paciente", "Nombre Paciente", "Edad", "Examen", "Rut", "Sexo",
                   "F.U.R.", "Fecha Nac.", "F.U.D.", "Hora FUD", "Fecha T.M.", "Hora T.M.", "VOL"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=6, column=col, value=header)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row = 7
        for codigo_paciente, paciente in pacientes.items():
            for examen in paciente.examenes:
                sheet.cell(row=row, column=1, value=codigo_paciente)
                sheet.cell(row=row, column=2, value=paciente.nombre.upper())
                sheet.cell(row=row, column=3, value=str(paciente.edad).upper())
                sheet.cell(row=row, column=4, value=examen.examen.upper())
                sheet.cell(row=row, column=5, value=paciente.rut.upper())
                sheet.cell(row=row, column=6, value=paciente.sexo.upper())
                sheet.cell(row=row, column=7, value="")
                sheet.cell(row=row, column=8, value=paciente.fecha_nacimiento.upper())
                sheet.cell(row=row, column=9, value="")
                sheet.cell(row=row, column=10, value="")
                sheet.cell(row=row, column=11, value=datetime.now().strftime('%d/%m/%Y'))
                sheet.cell(row=row, column=12, value=datetime.now().strftime('%H:%M'))
                sheet.cell(row=row, column=13, value='')
                row += 1

        workbook.save(archivo_excel)
        messagebox.showinfo("Éxito", f"Listado generado: {nombre_archivo}")

    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar archivo Excel: {str(e)}")
